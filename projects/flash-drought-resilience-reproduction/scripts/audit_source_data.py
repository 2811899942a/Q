#!/usr/bin/env python3
"""Audit the official Nature Source Data workbook without modifying it.

Usage:
    python scripts/audit_source_data.py path/to/41467_2026_70417_MOESM4_ESM.xlsx

Outputs a JSON report to stdout. Requires only openpyxl from the project environment.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    args = ap.parse_args()
    path = args.xlsx.resolve()

    wb = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    sheets = []
    total_formulas = 0
    for ws in wb.worksheets:
        nonempty = numeric = text = formulas = nan_text = 0
        headers = []
        for c in ws[1]:
            if c.value is not None:
                headers.append(str(c.value))
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                nonempty += 1
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    numeric += 1
                elif isinstance(v, str):
                    text += 1
                    if v.startswith("="):
                        formulas += 1
                    if v.strip().lower() == "nan":
                        nan_text += 1
        total_formulas += formulas
        sheets.append(
            {
                "sheet": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "nonempty": nonempty,
                "numeric": numeric,
                "text": text,
                "formula_cells": formulas,
                "nan_text_cells": nan_text,
                "headers": headers,
            }
        )

    report = {
        "file": path.name,
        "bytes": path.stat().st_size,
        "sha256": sha256(path),
        "sheet_count": len(wb.sheetnames),
        "formula_cells": total_formulas,
        "external_links": len(getattr(wb, "_external_links", [])),
        "defined_names": len(wb.defined_names),
        "sheets": sheets,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
