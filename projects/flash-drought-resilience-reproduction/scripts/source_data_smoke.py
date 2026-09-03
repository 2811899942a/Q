#!/usr/bin/env python3
"""Run low-cost numerical checks against the official Source Data workbook.

This does not reproduce the upstream climate processing. It verifies that the supplied
source values are internally consistent with several published rounded results.

Usage:
    python scripts/source_data_smoke.py path/to/41467_2026_70417_MOESM4_ESM.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


def numeric_column(ws, col_index: int):
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[col_index]
        if isinstance(v, (int, float)):
            out.append(float(v))
    return np.asarray(out, dtype=float)


def ols_slope(y: np.ndarray, start_year: int = 1950) -> float:
    x = np.arange(start_year, start_year + len(y), dtype=float)
    return float(np.polyfit(x, y, 1)[0])


def r2_sse(y: np.ndarray, pred: np.ndarray) -> float:
    return float(1.0 - np.sum((y - pred) ** 2) / np.sum((y - np.mean(y)) ** 2))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    args = ap.parse_args()
    wb = load_workbook(args.xlsx, read_only=True, data_only=True)

    # Main Fig. 1 time-series slopes.
    ws = wb["Figure1g-i"]
    slopes = {
        "count_anomaly_ols": ols_slope(numeric_column(ws, 1)),
        "severity_anomaly_ols": ols_slope(numeric_column(ws, 2)),
        "onset_speed_anomaly_ols": ols_slope(numeric_column(ws, 3)),
    }
    expected = {
        "count_anomaly_ols": 6.63,
        "severity_anomaly_ols": 0.012,
        "onset_speed_anomaly_ols": 0.018,
    }
    tolerances = {
        "count_anomaly_ols": 0.02,
        "severity_anomaly_ols": 0.001,
        "onset_speed_anomaly_ols": 0.001,
    }

    print("[Fig1 source-value OLS checks]")
    for key, value in slopes.items():
        ok = abs(value - expected[key]) <= tolerances[key]
        print(f"{key}: {value:.6f} -> {'PASS' if ok else 'FAIL'}")
        if not ok:
            raise SystemExit(2)

    # VIF final-factor screen.
    ws = wb["FigureS21a"]
    vif_values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        for v in row[2:6]:
            if isinstance(v, (int, float)):
                vif_values.append(float(v))
    max_vif = max(vif_values)
    print("\n[FigS21 VIF]")
    print(f"max final VIF = {max_vif:.6f}")
    if not max_vif < 5.0:
        raise SystemExit("FAIL: final predictor VIF >= 5")
    print("PASS: all final predictor VIF values are < 5")

    # RF observed/estimated pair diagnostics. These are diagnostics only; the exact
    # metric shown by the authors must be read from their MATLAB code.
    ws = wb["FigureS21b-e"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    pairs = [(1, 2), (3, 4), (5, 6), (7, 8)]
    labels = [
        "flash_hotspots",
        "slow_hotspots",
        "flash_non_hotspots",
        "slow_non_hotspots",
    ]
    print("\n[FigS21 observed/estimated diagnostics]")
    for label, (a, b) in zip(labels, pairs):
        xy = [(r[a], r[b]) for r in rows if isinstance(r[a], (int, float)) and isinstance(r[b], (int, float))]
        y = np.asarray([q[0] for q in xy], dtype=float)
        p = np.asarray([q[1] for q in xy], dtype=float)
        sse_r2 = r2_sse(y, p)
        corr2 = float(np.corrcoef(y, p)[0, 1] ** 2)
        print(f"{label}: n={len(y)}, SSE-R2={sse_r2:.4f}, corr^2={corr2:.4f}")

    print("\nSOURCE_DATA_SMOKE=PASS")


if __name__ == "__main__":
    main()
